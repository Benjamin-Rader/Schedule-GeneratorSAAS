import streamlit as st

st.title("🎈 Schedule Generator 🎈")

st.write("Upload your Excel file below. The app will process it and let you download the result.")

# File uploader for Excel files
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

output_bytes = None

if uploaded_file is not None:
    # Placeholder for your processing code
    # For now, just echo the uploaded file
    st.success("File uploaded successfully!")
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Font

    df = pd.read_excel(uploaded_file)
    ##df = openpyxl.load_workbook(r"C:\Users\Ben\Downloads\appointments (1).xlsx")
    print(df.head())

    df = df.drop(columns=["ID", "Alternate ID", 'Alternate ID 2', 'Email', 'Primary Major', 'College', 'Academic Year', 'Appointment Name', 'Appointment Reason', 'Created By', 'Created At', 'Status', 'Canceled By', 'Scheduled Duration', 'Comments', 'Attendee Count', 'Location', 'Location URL', 'Course Section', 'Appointment Type'])
    print(df.head())
    wb = openpyxl.load_workbook("RoomSheetTest2.xlsx")
    ws = wb.active  # Or workbook['SheetName']
    ws.title = 'RoomSheet'

    #### Today's Date Header
    date1 = ''
    date_str = df['Date'].iloc[1]  # e.g. '09/16/2025'
    ##gets the name of the month
    month_num = int(date_str.split('/')[0])  # Gets '09', converts to 9

    month_map = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December'
    }
    month_name = month_map.get(month_num, 'Unknown')
    ##
    ##Day of month
    dateday = df['Date'].iloc[1]
    if dateday == '1':
        th = 'st'
    elif dateday == '2':
        th = 'nd'
    elif dateday == '3':
        th = 'rd'
    else:
        th = 'th'

    dateday = dateday[3:5] + th
    ##
    ##year
    year = df['Date'].iloc[1]
    year = year[6:]
    ##
    date1 = date1 + df['Day'].iloc[1] + ', ' + month_name + ' ' + dateday + ' ' + year

    ws['A1'] = date1
    #####
    #####Names of SAs
    hours = {
        "eight": 0,
        "nine": 0,
        "ten": 0,
        "eleven": 0,
        "twelve": 0,
        "one": 0,
        "two": 0,
        "three": 0,
        "four": 0,
        "five": 0,
        "six": 0,
        "seven": 0
    }
    hour_map = {
        8: "eight",
        9: "nine",
        10: "ten",
        11: "eleven",
        12: "twelve",
        1: "one",
        2: "two",
        3: "three",
        4: "four",
        5: "five",
        6: "six",
        7: "seven"
    }

    for i in df['Scheduled Start Time']:
        # Split time and period (AM/PM)
        time_part, period = i.strip().split(' ')
        hour = int(time_part.split(':')[0])
        # Convert PM hours to 12-hour format
        if period == 'PM' and hour != 12:
            hour += 12
        if period == 'AM' and hour == 12:
            hour = 0

        # Map hour back to your dictionary keys
        # Only count hours between 1 and 12 (since your dictionary uses 12-hour keys)
        hour_key = None
        if 1 <= hour <= 12:
            hour_key = hour_map.get(hour)
        elif 13 <= hour <= 19:  # For PM hours like "7:00 PM" (which is 19:00)
            hour_key = hour_map.get(hour - 12)
        if hour_key:
            hours[hour_key] += 1

    label_map = {
        "eight": "8:00 AM",
        "nine": "9:00 AM",
        "ten": "10:00 AM",
        "eleven": "11:00 AM",
        "twelve": "12:00 PM",
        "one": "1:00 PM",
        "two": "2:00 PM",
        "three": "3:00 PM",
        "four": "4:00 PM",
        "five": "5:00 PM",
        "six": "6:00 PM",
        "seven": "7:00 PM"
    }

    for key in label_map:
        ws.append([label_map[key]])  # Section header
        for _ in range(hours[key]):
            ws.append([""])           # Blank row for each count

    # Save the workbook


    ### bold and fill time headings
    import re
    from openpyxl.styles import PatternFill, Alignment

    fsu_gold_fill = PatternFill(start_color="CEB888", end_color="CEB888", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    pattern = re.compile(r'^\d{1,2}:\d{2} [AP]M$')  # Matches 8:00 AM, 12:45 PM, etc.

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        found_time = False
        for cell in row:
            if isinstance(cell.value, str) and pattern.match(cell.value):
                found_time = True
                break
        if found_time:
            row_num = row[0].row
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            merged_cell = ws.cell(row=row_num, column=1)
            merged_cell.fill = fsu_gold_fill
            merged_cell.alignment = center_align
            merged_cell.font = Font(bold=True)

    ####change height of body rows to 39.75 px
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):  # Only column A
        cell = row[0]
        if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
            ws.row_dimensions[cell.row].height = 39.75

    ###inputs times to column A
    import openpyxl

    df_index = 0  # Start at the top of the DataFrame

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]  # Only column A
        if df_index >= len(df):
            break  # Stop if we've used all DataFrame values
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = df['Scheduled Start Time'].iloc[df_index]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                df_index += 1

    #############################################inputs names of SA################################################
    df_index = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]  # Only column B
        if df_index >= len(df):
            break  # Stop if we've used all DataFrame values
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                full_name = f"{df['First Name'].iloc[df_index]} {df['Last Name'].iloc[df_index]}"
                cell.value = full_name
                cell.alignment = Alignment(horizontal="left", vertical="center",wrap_text=True)
                df_index += 1

    ##############################adds sport#######################################################
    df_index = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
        cell = row[0]  # Only column C
        if df_index >= len(df):
            break  # Stop if we've used all DataFrame values
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = df['Group'].iloc[df_index]
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                df_index += 1

    ###########################################adds tutor/mentor####################################################
    df_index = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=4):
        cell = row[0]  # Only column C
        if df_index >= len(df):
            break  # Stop if we've used all DataFrame values
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = df['Host'].iloc[df_index]
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                df_index += 1

    ####################################course number or mentor###################################################
    df_index = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
        cell = row[0]  # Column E
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Fill only blank cells (None or empty string)
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                if df_index < len(df):
                    course_name = df['Course Name'].iloc[df_index]
                    # If the DataFrame value is nan, write "Mentoring"
                    if pd.isna(course_name):
                        cell.value = "Mentoring"
                    else:
                        cell.value = course_name  # Write the class code
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    df_index += 1  # Only increment when you assign a value
                else:
                    # DataFrame is exhausted, optionally keep filling with "Mentoring"
                    cell.value = "Mentoring"
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    ##Room assignment
    room_numbers = [3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18]

    # Step 1: Count sessions per host
    host_session_count = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        host_cell = row[3]  # Column D
        host_name = host_cell.value
        if host_name and isinstance(host_name, str) and host_name.strip():
            host_name = host_name.strip()
            host_session_count[host_name] = host_session_count.get(host_name, 0) + 1

    # Step 2: Assignment structures
    host_to_room = {}         # Host to room mapping
    room_in_use = {}          # Room -> host mapping (only if host has remaining sessions)

    # Step 3: Go through and assign rooms
    from openpyxl.styles import Alignment
    from datetime import datetime, timedelta

    room_numbers = [3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18]
    host_to_room = {}           # Host to room mapping
    room_last_end_time = {}     # Room to last session end time
    room_in_use = {}            # Room to current host
    used_rooms_this_cycle = set() # Rooms used in this cycle

    def parse_time(cell):
        if isinstance(cell.value, str):
            for fmt in ["%I:%M %p", "%H:%M"]:
                try:
                    return datetime.strptime(cell.value.strip(), fmt)
                except Exception:
                    continue
        elif isinstance(cell.value, datetime):
            return cell.value
        return None

    # Step 1: Count sessions per host
    host_session_count = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        host_cell = row[3]  # Column D
        host_name = host_cell.value
        if host_name and isinstance(host_name, str) and host_name.strip():
            host_name = host_name.strip()
            host_session_count[host_name] = host_session_count.get(host_name, 0) + 1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell_time = row[0]      # Column A (session time)
        cell_host = row[3]      # Column D (host name)
        cell_room = row[5]      # Column F (room assignment)
        if not isinstance(cell_room, openpyxl.cell.cell.MergedCell):
            if cell_room.value is None or (isinstance(cell_room.value, str) and cell_room.value.strip() == ""):
                host_name = cell_host.value
                session_start = parse_time(cell_time)
                if host_name and isinstance(host_name, str) and host_name.strip() and session_start:
                    host_name = host_name.strip()
                    # If host already has a room, use it
                    if host_name in host_to_room:
                        room = host_to_room[host_name]
                    else:
                        # Find available rooms (not currently in use)
                        available_rooms = []
                        for r in room_numbers:
                            last_end = room_last_end_time.get(r)
                            if (not last_end or last_end + timedelta(hours=1) <= session_start) and r not in room_in_use.values():
                                # Room is not occupied by someone with remaining sessions
                                if r not in used_rooms_this_cycle:
                                    available_rooms.append(r)
                        # If we've used all rooms in the cycle, reset
                        if not available_rooms:
                            used_rooms_this_cycle = set()
                            for r in room_numbers:
                                last_end = room_last_end_time.get(r)
                                if (not last_end or last_end + timedelta(hours=1) <= session_start) and r not in room_in_use.values():
                                    available_rooms.append(r)
                        if not available_rooms:
                            raise Exception("No available rooms for this session time!")
                        room = available_rooms[0]
                        host_to_room[host_name] = room
                        used_rooms_this_cycle.add(room)
                    cell_room.value = room
                    cell_room.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    # Update this room's last session end time (assuming 1-hour sessions)
                    room_last_end_time[room] = session_start
                    # Decrement host's session count
                    host_session_count[host_name] -= 1
                    if host_session_count[host_name] == 0:
                        # Host is done, room can be reassigned
                        if room in room_in_use:
                            del room_in_use[room]
                    else:
                        room_in_use[room] = host_name

    #########################################border the cells###################################################
    from openpyxl.styles import Border, Side

    black_side = Side(border_style="thin", color="000000")
    black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)

    # First, set borders for all regular cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = black_border

    # Now, explicitly set borders for all cells in merged ranges
    for merged_range in ws.merged_cells.ranges:
        for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row,
                                min_col=merged_range.min_col, max_col=merged_range.max_col):
            for cell in row:
                cell.border = black_border

    ######################replaces Womens Swimming with Swimming and Diving########################################
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell_c = row[2]  # Column C (0-based index)
        if cell_c.value == "Womens Swimming":
            cell_c.value = "Swimming and Diving"


    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    output_bytes = output.getvalue()
    

    st.write("Processing complete. Download your new Excel file below:")

if output_bytes:
    st.download_button(
        label="Download processed Excel file",
        data=output_bytes,
        file_name="output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
